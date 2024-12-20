# import pandas as pd
#
# data_frame = pd.read_excel('DataFiles/DataFile.xlsx')
# data_frame = data_frame.reset_index(drop=True)
#
# print(data_frame.head())
#
# new_row = {
#     'ID' : '3', 'Name':'Ghi'
# }
#
# data_frame.loc[len(data_frame)] = new_row
#
#
# print(data_frame.head())
#
# data_frame.to_excel('DataFiles/DataFile.xlsx')


import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('DataFiles/DataFile.xlsx')

# Get the sheet names
sheet_names = wb.sheetnames
print(f"Sheet Names: {sheet_names}")

# Get the active sheet
sheet = wb.active

# Read all rows and columns from the active sheet
for row in sheet.iter_rows(values_only=True):
    print(row)
